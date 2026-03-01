# -*- coding: utf-8 -*-
"""Registo de peças já existentes no teatro.app (XLSX).

Objectivo:
- Se a peça já existir no teatro.app, registar num .xlsx:
  título + link bilhetes + timestamp + url da página.
- Evitar duplicados (mesmo título + mesmo link).

Notas:
- Requer openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Tuple

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception as e:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    _OPENPYXL_ERR = e
else:
    _OPENPYXL_ERR = None


HEADERS = ["Data/Hora", "Peça", "Bilhetes", "URL teatro.app", "Observações"]


def _ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _normalizar(s: str) -> str:
    return (s or "").strip().lower()


def _key(title: str, ticket_url: str) -> Tuple[str, str]:
    return (_normalizar(title), _normalizar(ticket_url))


def _ensure_workbook(path: Path) -> None:
    if Workbook is None:
        raise RuntimeError(f"openpyxl não está disponível: {_OPENPYXL_ERR}")
    _ensure_parent_dir(path)
    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Existentes"
    ws.append(HEADERS)
    wb.save(path)


def append_exists_xlsx(
    path: Path,
    *,
    title: str,
    ticket_url: str,
    page_url: str = "",
    note: str = "Já existia no teatro.app",
) -> bool:
    """Apende registo ao XLSX se ainda não existir. Devolve True se adicionou."""
    _ensure_workbook(path)
    assert load_workbook is not None  # mypy
    wb = load_workbook(path)
    ws = wb.active

    wanted = _key(title, ticket_url)

    # procurar duplicado (ignora cabeçalho)
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[1] or "")
        u = str(row[2] or "")
        if _key(t, u) == wanted:
            return False

    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    ws.append([ts, title, ticket_url, page_url, note])
    wb.save(path)
    return True


def first_ticket_url_from_sessions_json(path: Path) -> str:
    """Extrai o primeiro ticket_url não-vazio do TEATROAPP_SESSIONS_JSON."""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return ""

    if not isinstance(data, list):
        return ""
    for item in data:
        if not isinstance(item, dict):
            continue
        u = str(item.get("ticket_url") or "").strip()
        if u:
            return u
    return ""
